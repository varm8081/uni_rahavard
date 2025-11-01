import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# -----------------------------
# تنظیمات اولیه
# -----------------------------
input_file = "دعبید_Merged.xlsx"         # فایل داده ورودی
output_file = "Abidi_Analysis.xlsx"  # خروجی با نام انگلیسی

# -----------------------------
# ۲۰ متغیر کلیدی با نام فارسی در داده
# -----------------------------
selected_vars = [
    "جمع دارایی‌های جاری",
    "جمع کل دارایی‌ها",
    "جمع بدهی‌های جاری",
    "جمع کل بدهی‌ها",
    "جمع حقوق صاحبان سهام",
    "جمع درآمدها",
    "سود (زیان) عملیاتی",
    "سود (زیان) ویژه پس از کسر مالیات",
    "سود و زیان انباشته در پایان دوره",
    "نسبت جاری",
    "نسبت آنی",
    "نسبت بدهی",
    "نسبت بدهی به ارزش ویژه",
    "بازده دارایی‌ها ROA",
    "بازدهی سرمایه ROE",
    "گردش موجودی کالا",
    "گردش دارایی‌های ثابت",
    "گردش مجموع دارایی‌ها",
    "سود ناخالص به فروش",
    "سود خالص به فروش"
]

# -----------------------------
# خواندن داده
# -----------------------------
df = pd.read_excel(input_file)

# فقط ردیف‌های مورد نظر را نگه می‌داریم
df = df[df.iloc[:, 0].isin(selected_vars)].copy()

# ستون اول (نام متغیر) را اندیس قرار می‌دهیم
df.set_index(df.columns[0], inplace=True)

# تبدیل همه داده‌ها به عدد (در صورت وجود کاراکتر غیرعددی)
df = df.applymap(lambda x: pd.to_numeric(str(x).replace(',', ''), errors='coerce'))

# -----------------------------
# 1️⃣ تحلیل توصیفی
# -----------------------------
desc_stats = df.T.describe().T  # توصیف آماری برای هر متغیر
desc_stats["Skewness"] = df.T.skew().values
desc_stats["Kurtosis"] = df.T.kurtosis().values

# نام ستون‌ها به انگلیسی برای خروجی
desc_stats.columns = [
    "Count", "Mean", "Std", "Min", "25%", "50%", "75%", "Max", "Skewness", "Kurtosis"
]

# -----------------------------
# 2️⃣ ماتریس همبستگی
# -----------------------------
corr_matrix = df.T.corr(method='pearson')

# -----------------------------
# ذخیره در فایل اکسل (دو شیت جدا)
# -----------------------------
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    desc_stats.to_excel(writer, sheet_name="Descriptive_Analysis")
    corr_matrix.to_excel(writer, sheet_name="Correlation_Matrix")

# -----------------------------
# رنگ‌بندی در اکسل برای دید بهتر
# -----------------------------
wb = load_workbook(output_file)

# رنگ‌بندی برای شیت تحلیل توصیفی (سبز: بالا، قرمز: پایین)
ws1 = wb["Descriptive_Analysis"]
for row in ws1.iter_rows(min_row=2, min_col=3):
    for cell in row:
        try:
            val = float(cell.value)
            if val < 0:
                cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")  # قرمز
            elif val > 0:
                cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")  # سبز
        except:
            pass

# رنگ‌بندی برای شیت همبستگی (قرمز منفی، آبی مثبت)
ws2 = wb["Correlation_Matrix"]
for row in ws2.iter_rows(min_row=2, min_col=2):
    for cell in row:
        try:
            val = float(cell.value)
            if val >= 0.7:
                cell.fill = PatternFill(start_color="B3C6FF", fill_type="solid")  # آبی پررنگ
            elif val <= -0.7:
                cell.fill = PatternFill(start_color="FF9999", fill_type="solid")  # قرمز پررنگ
            elif -0.3 < val < 0.3:
                cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")  # خاکستری روشن
        except:
            pass

wb.save(output_file)
print("✅ تحلیل توصیفی و همبستگی با موفقیت ذخیره شد →", output_file)
